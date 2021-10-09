"""
Classes for manage xlsx & csv files

"""
from openpyxl import load_workbook


class Excel:
    """
    Methods and properties about an Excel list of records
    In Property titles_coluns its found columns title
    In Properties cells_values (list) its found e cells value
    """
    __rows_max = 1000  # rows limit
    __columns_max = 20  # columns limit
    __columns_values = []  # values in a row
    cells_values = []  # List in witch embedded columns value
    title_columns = []  # columns Title
    records_with_data = 0
    columns_with_data = 0
    errors = ['OK',
              'il percorso del file è vuoto',
              'File non trovato',
              'Valore righe non amemsso',
              'Valore colonne non ammesso',
              'Il Nome sheet è vuoto']

    def __init__(self):
        """
        init properties
        """
        self.error = 0
        self.cells_values.clear()

    def load_excel_file(self,
                        full_path_file: str = '',
                        sheet_name: str = '',
                        first_row_of_data: int = 2,
                        first_column_of_data: int = 1,
                        rows_to_load: int = 1,
                        columns_to_load: int = 1,
                        row_title: int = 1) -> str:
        """
        Load file in Excel xlsx format and memo cells value
        :param sheet_name: Sheet name in file Excel
        :param full_path_file: Path & file name in file Excel
        :param first_row_of_data: First row to load with data. Default 1
        :param first_column_of_data: First column to load with data. Default 1
        :param rows_to_load: Number of rows to load. Its reduced if file has less rows.
        :param columns_to_load: Number of columns to load Its reduced if file has less columns.
        :param row_title: Row in witch found Columns intestation
        :return: 'OK' or error description string
        """

        # print('params:', full_path_file, sheet_name, first_row_of_data, first_column_of_data, rows_to_load,
        #      columns_to_load, row_title)

        # formal controls...
        if not full_path_file or not full_path_file.strip():
            self.error = 1
            return self.errors[1]

        if first_row_of_data > self.__rows_max or \
                rows_to_load > self.__rows_max or \
                first_row_of_data < 1 or \
                rows_to_load < 1:
            self.error = 1
            return self.errors[3]

        if first_column_of_data > self.__columns_max or \
                columns_to_load > self.__columns_max or \
                first_column_of_data < 1 or \
                columns_to_load < 1:
            self.error = 1
            return self.errors[4]

        if not sheet_name or not sheet_name.strip():
            self.error = 1
            return self.errors[5]

        # load file...
        self.error = 0

        try:
            workbook = load_workbook(filename=full_path_file)
            # sheet = workbook.active
            sheet = workbook[sheet_name]
            rows_max = sheet.max_row
            cols_max = sheet.max_column

            self.cells_values.clear()

            rows_to_load = min(rows_max, self.__rows_max, rows_to_load)
            self.records_with_data = rows_to_load
            columns_to_load = min(cols_max, self.__columns_max, columns_to_load)
            self.columns_with_data = columns_to_load

            # Load columns title...
            self.title_columns.clear()
            for c in range(first_column_of_data, columns_to_load+1):
                self.title_columns.append(sheet.cell(row=row_title, column=c).value)

            # Load matrix data...
            for r in range(first_row_of_data, rows_to_load+1):
                self.__columns_values.clear()
                self.__columns_values.append([])
                for c in range(first_column_of_data, columns_to_load):
                    # memo cell value...
                    self.__columns_values[0].append(sheet.cell(row=r, column=c).value)

                # print('Riga:', r)
                # print('columns values:', self.columns_values[0])
                self.cells_values.append(self.__columns_values[0])

            print("Celle:", self.cells_values)

            workbook.close()
            self.error = 0
        except Exception as e:
            print("Errore nella lettura del file: ", e)
            self.error = 1

        return self.error
