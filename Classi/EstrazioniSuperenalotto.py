"""
SUPERENALOTTO
-------------
Classe, metodi e proprietà per la gestione di una lista di estrazioni importate da foglio Excel

"""

from openpyxl import load_workbook
import Classi.EstrazioneSuperenalotto


class Estrazioni:
    """
    Methods and properties about an Excel list of Superenalotto extrations
    """

    def __init__(self):
        """
        EXTRACTIONS:
        number_extraction is List that contain number extraction
        number_extract is List that contain number extract
        First List is X axis in plot
        Second List is Y axis in plot
        """
        self.number_extraction = []
        self.number_extract = []
        self.extraction = Classi.EstrazioneSuperenalotto.Estrazione()
        self.error = 0
        self.begin_data_serie = 4

    def load_extraction_numbers_from_excel(self, full_file_name):
        """
        From specify path & file name, open file and load extraction numbers.
        The 5 numbers in extraction are set in single number whit EstrazioneSuperenalotto Class

        FORMAT FILE:
        ROWS 1, 2, 3 not significant
        ROW 4 begin date serie in format:
            Concorso	Data	N.1	N.2	N.3	N.4	N.5	N.6	Jolly	SuperStar

        :param full_file_name: path and file name. Format xlsx
        :return: retcode
        """
        try:
            workbook = load_workbook(filename=full_file_name)
            sheet = workbook.active
            self.number_extraction.clear()
            self.number_extract.clear()

            rows = sheet.max_row + 1
            for i in range(self.begin_data_serie, rows):
                # memo number extraction...
                self.number_extraction.append(sheet.cell(row=i, column=1).value)

                # memo numbers extract...
                numbers_list = [sheet.cell(row=i, column=3).value,
                                sheet.cell(row=i, column=4).value,
                                sheet.cell(row=i, column=5).value,
                                sheet.cell(row=i, column=6).value,
                                sheet.cell(row=i, column=7).value,
                                sheet.cell(row=i, column=8).value]
                string_extraction = self.extraction.send_numbers_about_single_extraction(numbers_list)  # transform in string
                number = self.extraction.numbers_extraction  # transform string in single number
                self.number_extract.append(number)

            # print("Estrazioni: ", self.number_extraction, self.number_extract)
            workbook.close()
            self.error = 0
        except Exception as e:
            print("Errore nella lettura del file: ", e)
            self.error = 1

        return self.error

    def load_extraction_first_number_from_excel(self, full_file_name):
        """
        From specify path & file name, open file and load extraction numbers.
        The first number in extraction are set in single number whit EstrazioneSuperenalotto Class
        :param full_file_name: path and file name. Format xlsx
        :return: retcode
        """
        try:
            self.load_extraction_numbers_from_excel(full_file_name)  # in memory all extractions

            # DA MODIFICARE--------------------

            for i in range(self.begin_data_serie, rows):
                # memo number extraction...
                self.number_extraction.append(sheet.cell(row=i, column=1).value)

                # read extraction numbers...
                numbers_list = [sheet.cell(row=i, column=3).value,
                                sheet.cell(row=i, column=4).value,
                                sheet.cell(row=i, column=5).value,
                                sheet.cell(row=i, column=6).value,
                                sheet.cell(row=i, column=7).value,
                                sheet.cell(row=i, column=8).value]
                string_extraction = self.extraction.send_numbers_about_single_extraction(numbers_list)
                number = self.extraction.number1
                self.number_extract.append(number)

            print("Estrazioni: ", self.number_extraction, self.number_extract)
            workbook.close()
            self.error = 0
        except Exception as e:
            print("Errore nella lettura del file: ", e)
            self.error = 1

        return self.error
